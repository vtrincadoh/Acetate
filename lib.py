import discogs_client
from discogs_client import models
import openpyxl
#from requests import get
#from shutil import copyfileobj

from const import SHOPIFY_DICT

discogs = discogs_client.Client('AltoqueInv/0.1', user_token='NYklFhDVlpWgnjwsXzkyvADMuqNKpDqbdPuxlPhN')

def cleanFilename(filename):
    split_filename = filename.split('\'')
    filename = split_filename[-2] if len(split_filename)>1 else filename
    return filename
def openBook(bookFilename):
    wb = openpyxl.load_workbook(bookFilename)
    return wb
def openSheet(wb):
    ws = wb.active
    return ws
def unpackIterRow(ws):
    last_row = ws.max_row
    iterable = ws.iter_rows(min_row=2, max_row=last_row, values_only=True)
    return iterable
def searchByProperty(data, property): #property es string, chequear valores de discogs_client
    def improvedMatch(releases):
        individual_release = next(releases, None)
        match individual_release:
            case None:
                return None
            case discogs_client.models.Release():
                return individual_release
            case discogs_client.models.Master():
                return individual_release.main_release
            case _:
                releases.remove(individual_release)
                return improvedMatch(releases)
    
    releases = discogs.search(str(data), type=property)
    return improvedMatch(iter(releases))
def extractTags(release, tag_attributes):
    def numberOfRecords(tracklist):
        tracks = {track.position[0] for track in tracklist}
        return len(tracks)
    
    '''
    Items in tag_attributes do not support list indexing, or sub-attributes
    '''
    tags = {}

    for attr in tag_attributes:
        item_in_dict = {attr: getattr(release, attr)}
        tags.update(item_in_dict)

        match attr:
            case ('artists' | 'labels'):
                tags[attr] = tags[attr][0].name.upper()
            case 'images':
                tags[attr] = tags[attr][0]['resource_url']
            case 'genres':
                tags[attr] = [tag.upper() for tag in tags[attr]]
                tags[attr] = str(tags[attr])
            case 'tracklist':
                tags[attr] = numberOfRecords(tags[attr])
            case _:
               tags[attr] = str(tags[attr]).upper()
    return tags
'''def downloadImg(image_url, filename): # Toma una url y descarga la imagen. Devuelve bool si fue exitoso.
    filename = str(filename) + '.' + image_url.split(".")[-1]
    r = get(image_url, stream = True)
    if r.status_code == 200:
        r.raw.decode_content = True
        with open(filename,'wb') as f:
            copyfileobj(r.raw, f)
        return True
    else:
        return False'''
def assignTags(tags, worksheet):
    worksheet.append(tags)
    return worksheet
def saveSheet(workbook):
    workbook.save(filename= 'ouput.xlsx')
    return
def valueOfShopifyProperty(tags_row, shopify_key):
    def reqTags(shopify_attr):
        req_props = SHOPIFY_DICT[shopify_attr]
        req_tags = req_props['args']
        req_function = req_props['func']
        return req_tags, req_function

    req_tags, req_function = reqTags(shopify_key)
    args = (tags_row[tag] for tag in req_tags)
    shopify_prop = req_function(*args) 
    return shopify_prop 